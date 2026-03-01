from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook

from src.models.schemas import Category, Geography, Source

logger = logging.getLogger(__name__)

SOURCES_FILE = Path(__file__).resolve().parents[2] / ".docs" / "Sources.xlsx"

SCOPE_TO_CATEGORY: dict[str, Category] = {
    "political": Category.POLITICS,
    "economics": Category.ECONOMICS,
    "social trends": Category.SOCIAL,
    "tech": Category.TECH,
}

GEO_TO_GEOGRAPHY: dict[str, Geography] = {
    "portugal": Geography.PORTUGAL,
}


def _normalize_url(raw_url: str) -> str:
    url = raw_url.strip()
    if not url.startswith("http"):
        url = f"https://www.{url}"
    return url


def _parse_geography(raw_geo: str) -> Geography:
    return GEO_TO_GEOGRAPHY.get(raw_geo.strip().lower(), Geography.GLOBAL)


def _infer_language(geography: Geography) -> str:
    return "pt" if geography == Geography.PORTUGAL else "en"


def load_sources_from_excel(path: Path = SOURCES_FILE) -> list[Source]:
    """Read sources from the Excel file, one Source per row."""
    if not path.exists():
        logger.error("❌ Sources file not found: %s", path)
        return []

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    sources: list[Source] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        scope, name, description, link, geo_raw = row[:5]

        if not name or not link:
            logger.warning("⚠️ Skipping row %d: missing name or link", i)
            continue

        category = SCOPE_TO_CATEGORY.get(str(scope).strip().lower())
        if not category:
            logger.warning("⚠️ Skipping row %d ('%s'): unknown scope '%s'", i, name, scope)
            continue

        geography = _parse_geography(str(geo_raw or "World"))
        language = _infer_language(geography)

        sources.append(
            Source(
                name=str(name).strip(),
                url=_normalize_url(str(link)),
                category=category,
                geography=geography,
                language=language,
                research_hint=str(description or "").strip(),
            )
        )

    wb.close()
    logger.info("📋 Loaded %d sources from %s", len(sources), path.name)
    return sources


def get_all_sources() -> list[Source]:
    return load_sources_from_excel()


def get_sources_by_category(category: Category) -> list[Source]:
    return [s for s in get_all_sources() if s.category == category]
